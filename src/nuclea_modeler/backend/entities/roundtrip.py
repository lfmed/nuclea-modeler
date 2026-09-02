"""Round-trip de edição via CSV (v1.0035, feedback do cliente).

Fluxo pedido pelo cliente:
  exportar dados (tabela/coluna/esquema) filtrados → editar o arquivo fora do app
  → reimportar → o app mostra o DIFF e submete à APROVAÇÃO.

Decisão de design: em vez de reaproveitar o diff-engine da extração (que só
compara tipo/PK/comentários e NÃO logical_name/descrição — justamente o que o
cliente quer ajustar), montamos um **ticket editorial** no MESMO formato do
editor manual (session ticket com `field_changes` `attribute:NOME.update` para
colunas e campos de entidade em ENTITY_FIELDS). Esse caminho de apply já é
comprovado e cobre logical_name, description_md, domínio, criticidade, tipo, PK,
nullable e comentário. Assim o "diff" aparece no ticket e a aprovação usa o fluxo
existente (approve/apply), sem código de materialização novo.

Formato do CSV (uma linha por COLUNA; grão coluna cobre tabela/coluna/esquema):

  schema, table, table_logical, table_description, table_domain, table_criticality,
  column, column_logical, data_type, is_pk, is_nullable, column_description

- A CHAVE é (schema, table, column). Linhas cujo (schema.table.column) não existe
  no catálogo viram `attribute_add`. Colunas do catálogo ausentes no CSV NÃO são
  removidas (import é aditivo/editorial — remoção some via UI, não por omissão no
  arquivo, evitando exclusões acidentais).
- Campos vazios no CSV = "não mexer" (mantém o valor atual do catálogo).
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from ..core import delta
from ..core._nuclea_config import get_settings
from ..tickets.session import get_or_create_session_ticket, stage_entity_change

# round 6 pt 22 — o cliente embute a classificação LGPD na própria descrição, como
# sufixo `… | CLASSIFICACAO=LGPD_*` (nos DDLs COMMENT ON, no CSV e no .xlsx do
# Embarcadero). No import de metadados, esse token vira uma FLAG LGPD aplicada ao
# atributo (que já existe → tem attribute_id real) e é REMOVIDO da descrição.
_CLASSIFICACAO_TO_FLAG = {
    "LGPD_PESSOAL": "dados-pessoais",
    "LGPD_SENSIVEL": "dados-sensiveis",
    "LGPD_IDENTIFICAVEL": "titular-identificado",
}
_FLAG_TO_CLASSIFICACAO = {v: k for k, v in _CLASSIFICACAO_TO_FLAG.items()}
_CLASSIF_RE = re.compile(r"\s*\|\s*CLASSIFICACAO\s*=\s*([A-Za-z_]+)\s*$")


def _split_classificacao(text: str | None) -> tuple[str | None, str | None]:
    """Separa a descrição do sufixo `| CLASSIFICACAO=LGPD_*`.

    Devolve ``(descricao_limpa, flag_key | None)``. Se não houver token, devolve o
    texto original e ``None``. Token desconhecido → mantém o texto e não flageia.
    """
    if not text:
        return text, None
    m = _CLASSIF_RE.search(text)
    if not m:
        return text, None
    flag_key = _CLASSIFICACAO_TO_FLAG.get(m.group(1).upper())
    if not flag_key:
        return text, None  # token desconhecido: não mexe na descrição
    clean = _CLASSIF_RE.sub("", text).strip() or None
    return clean, flag_key


def _apply_classificacao_flags(sql, actor: str, intents: list[tuple[str, str]]) -> int:
    """Aplica as flags LGPD derivadas do CLASSIFICACAO nos atributos existentes.

    `intents`: lista de ``(attribute_id, flag_key)``. Idempotente e com propagação
    LGPD para a entidade — reusa o core de flags (não duplica a lógica). Flags são
    aplicadas DIRETAMENTE (não são editoriais/staged, como no resto do app).
    """
    if not intents:
        return 0
    # Import tardio: evita qualquer ordem de import no boot (flags.router não
    # importa entities, então não há ciclo, mas mantém o módulo leve).
    from ..flags.router import _apply_attribute_flag_core, _fetch_flag
    s = get_settings()
    flag_cache: dict[str, Any] = {}
    applied = 0
    for attribute_id, flag_key in intents:
        flag = flag_cache.get(flag_key, "MISS")
        if flag == "MISS":
            row = delta.fetch_one_params(
                sql,
                f"SELECT flag_id FROM {s.fq_table('flags')} WHERE flag_key = :k",
                [delta.param("k", flag_key)],
            )
            flag = _fetch_flag(sql, row[0]) if row else None
            flag_cache[flag_key] = flag
        if not flag:
            continue
        try:
            _apply_attribute_flag_core(
                sql, attribute_id=attribute_id, flag=flag,
                justification="Importado de CLASSIFICACAO (metadados)", actor=actor,
            )
            applied += 1
        except Exception:  # noqa: BLE001 — uma flag não pode derrubar o import
            continue
    return applied

# Cabeçalho canônico do arquivo de round-trip. A ordem é estável (o parser aceita
# qualquer ordem desde que os nomes batam — usa DictReader).
CSV_HEADERS = [
    "schema", "table", "table_logical", "table_description",
    "table_domain", "table_criticality",
    "column", "column_logical", "data_type", "is_pk", "is_nullable",
    "column_description",
]


def _b(v: Any) -> str:
    return "" if v is None else str(v)


def export_system_csv(sql, system_id: str) -> str:
    """Exporta todas as colunas do sistema (com contexto de tabela/esquema) como
    CSV re-importável. Uma linha por coluna."""
    s = get_settings()
    rows = delta.fetch_all_params(
        sql,
        f"""
        SELECT e.schema_name, e.technical_name, e.logical_name, e.description_md,
               e.domain, e.criticality,
               a.technical_name, a.logical_name, a.native_data_type,
               a.is_primary_key, a.is_nullable, a.description_md,
               a.ordinal_position
        FROM {s.fq_table('attributes')} a
        JOIN {s.fq_table('entities')} e ON e.entity_id = a.entity_id
        WHERE e.system_id = :sid
        ORDER BY e.schema_name, e.technical_name,
                 COALESCE(a.ordinal_position, 999999), a.technical_name
        """,
        [delta.param("sid", system_id)],
    )
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(CSV_HEADERS)
    for r in rows:
        w.writerow([
            _b(r[0]), _b(r[1]), _b(r[2]), _b(r[3]), _b(r[4]), _b(r[5]),
            _b(r[6]), _b(r[7]), _b(r[8]),
            "true" if delta.as_bool(r[9]) else "false",
            "true" if (r[10] is None or delta.as_bool(r[10])) else "false",
            _b(r[11]),
        ])
    return buf.getvalue()


def _load_catalog(sql, system_id: str) -> dict[str, dict]:
    """Mapa (schema.table) -> {entity_id, entity fields, attrs: {col: {...}}}."""
    s = get_settings()
    rows = delta.fetch_all_params(
        sql,
        f"""
        SELECT e.entity_id, e.schema_name, e.technical_name, e.logical_name,
               e.description_md, e.domain, e.criticality, e.entity_type,
               a.technical_name, a.logical_name, a.native_data_type,
               a.is_primary_key, a.is_nullable, a.description_md, a.ordinal_position,
               a.attribute_id
        FROM {s.fq_table('entities')} e
        LEFT JOIN {s.fq_table('attributes')} a ON a.entity_id = e.entity_id
        WHERE e.system_id = :sid
        """,
        [delta.param("sid", system_id)],
    )
    cat: dict[str, dict] = {}
    for r in rows:
        key = f"{r[1]}.{r[2]}"
        ent = cat.setdefault(key, {
            "entity_id": r[0], "schema_name": r[1], "technical_name": r[2],
            "logical_name": r[3], "description_md": r[4], "domain": r[5],
            "criticality": r[6], "entity_type": r[7] or "TABLE", "attrs": {},
        })
        if r[8]:  # tem coluna
            ent["attrs"][r[8]] = {
                "logical_name": r[9], "native_data_type": r[10],
                "is_primary_key": delta.as_bool(r[11]),
                "is_nullable": (r[12] is None or delta.as_bool(r[12])),
                "description_md": r[13], "ordinal_position": r[14],
                "attribute_id": r[15],  # round 6 pt 22 — p/ aplicar flag CLASSIFICACAO
            }
    return cat


def _norm(v: str | None) -> str | None:
    """Célula vazia do CSV = None ("não mexer")."""
    if v is None:
        return None
    v = v.strip()
    return v or None


def parse_and_stage_csv(sql, actor: str, system_id: str, csv_text: str) -> dict:
    """Compara o CSV com o catálogo e monta UM ticket editorial (session ticket).

    Retorna um resumo com contagens + ticket_id (ou None se nada mudou). NÃO
    aplica nada — só stage (aparece como diff pendente de aprovação).
    """
    cat = _load_catalog(sql, system_id)
    reader = csv.DictReader(io.StringIO(csv_text))
    missing = [h for h in ("schema", "table", "column") if h not in (reader.fieldnames or [])]
    if missing:
        raise ValueError(f"CSV sem colunas obrigatórias: {', '.join(missing)}")

    # Acumula por entidade: field_changes + payload de entidade.
    per_entity: dict[str, dict] = {}
    unknown_tables: set[str] = set()
    # round 6 pt 22: (attribute_id, flag_key) das colunas existentes cuja descrição
    # trazia `| CLASSIFICACAO=…`. Aplicadas DIRETAMENTE ao fim (flags não são staged).
    flag_intents: list[tuple[str, str]] = []

    for raw in reader:
        schema = _norm(raw.get("schema"))
        table = _norm(raw.get("table"))
        column = _norm(raw.get("column"))
        if not schema or not table or not column:
            continue
        key = f"{schema}.{table}"
        cat_ent = cat.get(key)
        if not cat_ent:
            # Tabela não existe no catálogo — round-trip é para AJUSTE de dados
            # existentes; criar tabela nova fica para o import de DDL. Registra e ignora.
            unknown_tables.add(key)
            continue

        acc = per_entity.setdefault(key, {
            "entity_id": cat_ent["entity_id"],
            "schema_name": schema, "technical_name": table,
            "entity_type": cat_ent["entity_type"],
            "ent_changes": {}, "attr_changes": [], "attr_adds": [],
        })

        # --- Entity-level (uma vez por tabela; usa a 1ª linha que trouxer valor) ---
        for csv_col, ent_field in (
            ("table_logical", "logical_name"),
            ("table_description", "description_md"),
            ("table_domain", "domain"),
            ("table_criticality", "criticality"),
        ):
            val = _norm(raw.get(csv_col))
            if val is not None and val != _b(cat_ent.get(ent_field)) and ent_field not in acc["ent_changes"]:
                acc["ent_changes"][ent_field] = val

        # --- Column-level ---
        cat_a = cat_ent["attrs"].get(column)
        col_logical = _norm(raw.get("column_logical"))
        data_type = _norm(raw.get("data_type"))
        # round 6 pt 22: destaca o token CLASSIFICACAO → flag; a descrição staged
        # fica LIMPA (sem o sufixo). A flag só se aplica a coluna JÁ existente.
        col_desc, _classif_key = _split_classificacao(_norm(raw.get("column_description")))
        if cat_a is not None and _classif_key and cat_a.get("attribute_id"):
            flag_intents.append((cat_a["attribute_id"], _classif_key))
        is_pk_raw = _norm(raw.get("is_pk"))
        is_null_raw = _norm(raw.get("is_nullable"))
        is_pk = None if is_pk_raw is None else is_pk_raw.lower() in ("true", "1", "sim", "yes")
        is_null = None if is_null_raw is None else is_null_raw.lower() in ("true", "1", "sim", "yes")

        if cat_a is None:
            # coluna nova → attribute_add (payload completo)
            acc["attr_adds"].append({
                "technical_name": column,
                "logical_name": col_logical,
                "native_data_type": data_type,
                "is_primary_key": bool(is_pk),
                "is_nullable": True if is_null is None else is_null,
                "description_md": col_desc,
            })
        else:
            # coluna existente → só stage se algo mudou (payload COMPLETO no
            # formato attribute:NOME.update, que o apply trata).
            changed = (
                (col_logical is not None and col_logical != _b(cat_a["logical_name"])) or
                (data_type is not None and data_type != _b(cat_a["native_data_type"])) or
                (col_desc is not None and col_desc != _b(cat_a["description_md"])) or
                (is_pk is not None and is_pk != bool(cat_a["is_primary_key"])) or
                (is_null is not None and is_null != bool(cat_a["is_nullable"]))
            )
            if changed:
                acc["attr_changes"].append({
                    "technical_name": column,
                    # valores: CSV quando informado, senão mantém o catálogo
                    "logical_name": col_logical if col_logical is not None else cat_a["logical_name"],
                    "native_data_type": data_type if data_type is not None else cat_a["native_data_type"],
                    "is_primary_key": is_pk if is_pk is not None else bool(cat_a["is_primary_key"]),
                    "is_nullable": is_null if is_null is not None else bool(cat_a["is_nullable"]),
                    "description_md": col_desc if col_desc is not None else cat_a["description_md"],
                    "ordinal_position": cat_a["ordinal_position"],
                })

    # --- Monta as entries do ticket (formato editorial) ---
    entries: list[dict] = []
    n_ent = n_attr = 0
    for key, acc in per_entity.items():
        field_changes: list[dict] = []
        for fld, val in acc["ent_changes"].items():
            field_changes.append({"field": fld, "before": None, "after": val})
        for a in acc["attr_adds"]:
            field_changes.append({
                "field": f"attribute_add:{a['technical_name']}", "before": None, "after": a,
            })
            n_attr += 1
        for a in acc["attr_changes"]:
            field_changes.append({
                "field": f"attribute:{a['technical_name']}.update", "before": None, "after": a,
            })
            n_attr += 1
        if not field_changes:
            continue
        n_ent += 1
        entries.append({
            "op": "change",
            "schema_name": acc["schema_name"],
            "technical_name": acc["technical_name"],
            "entity_type": acc["entity_type"],
            "payload": {
                "target_entity_id": acc["entity_id"],
                **acc["ent_changes"],
            },
            "field_changes": field_changes,
        })

    # Flags LGPD do CLASSIFICACAO aplicadas DIRETAMENTE (idempotente) — independem
    # de haver mudança de descrição/tipo (a coluna pode já ter a descrição certa e
    # faltar só a flag). Deduplica (attribute_id, flag_key).
    flags_applied = _apply_classificacao_flags(sql, actor, sorted(set(flag_intents)))

    if not entries:
        msg = (
            f"{flags_applied} flag(s) LGPD aplicada(s); nenhuma mudança de descrição/tipo."
            if flags_applied else "Nenhuma mudança detectada em relação ao catálogo."
        )
        return {
            "ticket_id": None, "entities_changed": 0, "columns_changed": 0,
            "flags_applied": flags_applied,
            "unknown_tables": sorted(unknown_tables),
            "message": msg,
        }

    ticket_id, diff = get_or_create_session_ticket(
        sql, actor, system_id, title_hint="Importação de metadados (round-trip)"
    )
    for entry in entries:
        diff = stage_entity_change(sql, ticket_id, diff, entry)

    flag_msg = f" · {flags_applied} flag(s) LGPD aplicada(s)" if flags_applied else ""
    return {
        "ticket_id": ticket_id,
        "entities_changed": n_ent,
        "columns_changed": n_attr,
        "flags_applied": flags_applied,
        "unknown_tables": sorted(unknown_tables),
        "message": f"{n_ent} tabela(s) e {n_attr} coluna(s) com ajustes — revise e aprove no ticket.{flag_msg}",
    }


# ─── xlsx do Embarcadero (round 6 pt 22) ──────────────────────────────────────
#
# Formato do cliente (descricoes_embarcadero.xlsx): sheet única, header
#   table | table_description | column | column_description
# — sem coluna de schema. No import resolvemos o schema pelo catálogo (nome de
# tabela único) e convertemos para o CSV canônico, reusando TODO o parse_and_stage_csv
# (staging editorial + CLASSIFICACAO→flag). No export geramos o mesmo layout.


def _read_xlsx_rows(xlsx_bytes: bytes) -> list[dict[str, str | None]]:
    """Lê o .xlsx (header table/table_description/column/column_description) em
    dicts. Aceita qualquer ordem de coluna. openpyxl é dependência (server-side);
    erro amigável se ausente."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "Suporte a .xlsx indisponível no servidor (openpyxl não instalado)."
        ) from exc
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = [(str(c).strip().lower() if c is not None else "") for c in next(it)]
    except StopIteration:
        wb.close()
        return []
    idx = {name: i for i, name in enumerate(header)}

    def cell(row: tuple, name: str) -> str | None:
        i = idx.get(name)
        if i is None or i >= len(row) or row[i] is None:
            return None
        return str(row[i])

    out: list[dict[str, str | None]] = []
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        out.append({
            "table": cell(row, "table"),
            "table_description": cell(row, "table_description"),
            "column": cell(row, "column"),
            "column_description": cell(row, "column_description"),
        })
    wb.close()
    return out


def _table_to_schema(cat: dict[str, dict]) -> dict[str, str | None]:
    """Mapa nome_de_tabela → schema, ou None quando ambíguo (2+ schemas)."""
    seen: dict[str, str | None] = {}
    for ent in cat.values():
        t = ent["technical_name"]
        seen[t] = ent["schema_name"] if t not in seen else None
    return seen


def parse_and_stage_xlsx(sql, actor: str, system_id: str, xlsx_bytes: bytes) -> dict:
    """Importa metadados do .xlsx do Embarcadero (pt 22).

    O .xlsx NÃO traz schema; resolvemos pelo catálogo (nome de tabela único),
    convertemos para o CSV canônico e delegamos a `parse_and_stage_csv` — assim o
    staging editorial + CLASSIFICACAO→flag são exatamente os mesmos do CSV.
    """
    xlsx_rows = _read_xlsx_rows(xlsx_bytes)
    if not xlsx_rows:
        return {"ticket_id": None, "entities_changed": 0, "columns_changed": 0,
                "flags_applied": 0, "unknown_tables": [],
                "message": "Planilha vazia ou sem cabeçalho reconhecido (table/column)."}
    tbl_to_schema = _table_to_schema(_load_catalog(sql, system_id))
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(CSV_HEADERS)
    for r in xlsx_rows:
        table = (r.get("table") or "").strip()
        column = (r.get("column") or "").strip()
        if not table or not column:
            continue
        schema = tbl_to_schema.get(table) or ""
        w.writerow([
            schema, table, "", r.get("table_description") or "", "", "",
            column, "", "", "", "", r.get("column_description") or "",
        ])
    return parse_and_stage_csv(sql, actor, system_id, buf.getvalue())


def export_system_xlsx(sql, system_id: str) -> bytes:
    """Exporta metadados no formato .xlsx do Embarcadero (pt 22): 4 colunas
    table|table_description|column|column_description. Reconstrói o sufixo
    `| CLASSIFICACAO=…` a partir das flags LGPD do atributo (round-trip fiel)."""
    try:
        from openpyxl import Workbook
    except Exception as exc:  # noqa: BLE001
        raise ValueError(
            "Suporte a .xlsx indisponível no servidor (openpyxl não instalado)."
        ) from exc
    s = get_settings()
    rows = delta.fetch_all_params(
        sql,
        f"""
        SELECT e.technical_name, e.description_md, a.technical_name, a.description_md,
               a.ordinal_position, a.attribute_id
        FROM {s.fq_table('attributes')} a
        JOIN {s.fq_table('entities')} e ON e.entity_id = a.entity_id
        WHERE e.system_id = :sid
        ORDER BY e.schema_name, e.technical_name,
                 COALESCE(a.ordinal_position, 999999), a.technical_name
        """,
        [delta.param("sid", system_id)],
    )
    flag_rows = delta.fetch_all_params(
        sql,
        f"""
        SELECT af.attribute_id, f.flag_key
        FROM {s.fq_table('attribute_flags')} af
        JOIN {s.fq_table('attributes')} a ON a.attribute_id = af.attribute_id
        JOIN {s.fq_table('entities')} e ON e.entity_id = a.entity_id
        JOIN {s.fq_table('flags')} f ON f.flag_id = af.flag_id
        WHERE e.system_id = :sid
        """,
        [delta.param("sid", system_id)],
    )
    classif_by_attr: dict[str, str] = {}
    for aid, fkey in flag_rows:
        token = _FLAG_TO_CLASSIFICACAO.get(fkey)
        if token and aid not in classif_by_attr:
            classif_by_attr[aid] = token

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws.append(["table", "table_description", "column", "column_description"])
    for r in rows:
        desc = r[3] or ""
        token = classif_by_attr.get(r[5])
        if token:
            desc = (desc + f" | CLASSIFICACAO={token}").strip()
        ws.append([_b(r[0]), _b(r[1]), _b(r[2]), desc])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
